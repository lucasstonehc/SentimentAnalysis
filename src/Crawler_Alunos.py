#Programa que coleta todos os twittes já realizados de determinado usuário
#Os dados são colocados em uma planilha no formato xlsx.

#Adaptação dos códigos de Kaorw e Paulo:

#https://gist.github.com/Kaorw/7594044 
#'Grap multiple user's user_timeline from twitter API and save to Excel'

#https://paulovasconcellos.com.br/aprenda-a-fazer-um-analisador-de-sentimentos-do-twitter-em-python-3979454f2d0d
#'Aprenda a fazer um Analisador de Sentimentos do Twitter em Python'


#Biblioteca de escrita na planilha
import xlsxwriter

#Biblioteca de dados do Twitter
import tweepy 
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import re

#Dados de conta
consumer_key ='5W7W6BSMHMsEtnLNNUJSo4jcs'
consumer_secret ='XJi5T7Ta5Co18Hbfgh1HvNubZ9pccx8y8q5uqUto5h8Sk9swo6'

access_key = '954784886157643778-buXhPhSb4eysyJVq01xTymPpzJBFO6a'
access_secret = 'yIFSSTBhzx1ZblAXa9ecrWKHuEtrZ2KJhur7Ddecf750G'


#Funcao que coleta todos os tweets
def get_all_tweets(screen_name):

    auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
    auth.set_access_token(access_key, access_secret)
    api = tweepy.API(auth,wait_on_rate_limit=True)

    alltweets = []  
    new_tweets = []
    outtweets = []

    new_tweets = api.user_timeline(screen_name = screen_name,count=200,retweet=False,full_text= True)

    alltweets.extend(new_tweets)

	#save the id of the oldest tweet less one
    oldest = alltweets[-1].id - 1

    #keep grabbing tweets until there are no tweets left to grab
    while len(new_tweets) > 0:
        #print "getting tweets before %s" % (oldest)

        #all subsiquent requests use the max_id param to prevent duplicates
        new_tweets = api.user_timeline(screen_name = screen_name,count=200,max_id=oldest,retweet=False,full_text= True)

        #save most recent tweets
        alltweets.extend(new_tweets)

        #update the id of the oldest tweet less one
        oldest = alltweets[-1].id - 1

        #print "...%s tweets downloaded so far" % (len(alltweets))

    #transform the tweepy tweets into a 2D array
    outtweets = [[tweet.id_str, tweet.created_at,tweet.coordinates,tweet.geo,tweet.truncated,tweet.text] for tweet in alltweets]

    return outtweets

def seasons_of_tweet(created_at):
	#get attr created_at from tweet object
	data = created_at # get when has cretead

	#strftime is a function from library´s datetime
	month = data.strftime(r"%b") #get month from data
	day = data.strftime(r"%d") # get day from data
	season = " "
	#SEASONS IN BRAZIL 
	#Summer > [December January February]
	#started: 21 December  Finished: 19 March

	#Autumn > [March April May]
	#started: 20 March  Finished: 19 June

	#Winter > [June July August]
	#started: 20 June Finished: 21 September

	#Spring > [September October November]
	#started: 22 September  Finished: 20 December

	if month == 'Dec' and day >= '21':
		season = 'summer'
	elif month == 'Jan':
		season = 'summer'
	elif month == 'Feb':
		season = 'summer'
	elif month == 'Mar' and day <= '19':
		season = 'summer'
	elif month == 'Mar' and day >= '20':
		season = 'Autumn'
	elif month == 'Apr':
		season = 'Autumn'
	elif month == 'May':
		season = 'Autumn'
	elif month == 'Jun' and day <= '19':
		season = 'Autumn'
	elif month == 'Jun' and day >= '20':
		season = 'Winter'
	elif month == 'Jul':
		season = 'Winter'
	elif month == 'Aug':
		season = 'Winter'
	elif month == 'Sep' and day <= '21':
		season = 'Winter'
	elif month == 'Sep' and day >= '22':
		season = 'Winter'
	elif month == 'Oct':
		season = 'Spring'
	elif month == 'Nov':
		season = 'Spring'
	elif month == 'Dec' and day <= '20':
		season = 'Spring'

	return season

def clear_twitter(line):
	line = re.sub(r'https:\S+', '',line)
	line = re.sub(r'RT @\S+', '',line)
	line = re.sub(r';\S+', '',line)
	line = re.sub(r'@\S+', '',line)
	line = re.sub(r'3\S+', '',line)
	line = re.sub(r'zap\S+', '',line)

	return line

def write_worksheet(id_str):

	#formating for excel
	format01 = workbook.add_format()
	format02 = workbook.add_format()
	format03 = workbook.add_format()
	format04 = workbook.add_format()
	format01.set_align('center')
	format01.set_align('vcenter')
	format02.set_align('center')
	format02.set_align('vcenter')
	format03.set_align('center')
	format03.set_align('vcenter')
	format03.set_bold()
	format04.set_align('vcenter')
	format04.set_text_wrap()

	out1 = []
	header = ["id","created_at","coordinates-x","coordinates-y","truncated","text","season"]

	worksheet = workbook.add_worksheet(id_str)

	out1 = get_all_tweets(id_str)  #all tweets from that person
	row = 0
	col = 0

	worksheet.set_column('A:A', 20)
	worksheet.set_column('B:B', 18)
	worksheet.set_column('C:C', 13)
	worksheet.set_column('D:D', 13)
	worksheet.set_column('E:E', 20)
	worksheet.set_column('F:F', 120)

	for h_item in header:
		worksheet.write(row, col, h_item, format03)
		col = col + 1

	row += 1
	col = 0
	
	for o_item in out1: #for to get each tweeter
		write = []
		cord1 = 0
		cord2 = 0
		write = [o_item[0], o_item[1], o_item[4], o_item[5]]

		if o_item[2]:
			cord1 = o_item[2]['coordinates'][0]
			cord2 = o_item[2]['coordinates'][1]
		else:
			cord1 = ""
			cord2 = ""

		format01.set_num_format('yyyy/mm/dd hh:mm:ss')
		worksheet.write(row, 0, write[0], format02)
		worksheet.write(row, 1, write[1], format01)
		worksheet.write(row, 2, cord1, format02)
		worksheet.write(row, 3, cord2, format02)
		worksheet.write(row, 4, write[2], format02)
		line = clear_twitter(write[3]) #clear the twitter 
		worksheet.write(row, 5,line, format04) #this column had text
		season = seasons_of_tweet(write[1]) #set which season it is.
		worksheet.write(row, 6, season, format04)
		row += 1
		col = 0


workbook = xlsxwriter.Workbook('alltwittersprofile.xlsx')

#to read all twitters profile from file
wbLoaded = load_workbook("ansiedade_crawler_IFs.xlsx")
sheet = wbLoaded.active
for row in range(1,100,1):
	profile_name = str(sheet.cell(row=row, column=1).value)
	try:
		write_worksheet(profile_name)
	except:
		print("Twitter Profile has not found!")

wbLoaded.close()
workbook.close()


